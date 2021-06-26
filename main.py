import os
import scipy.io
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt

from algorithm import algorithm
from analysis import (calculate_average_and_std_of_incline, calculate_difference_in_response_to_treatment, print_graph_of_average_and_std_of_incline)

if __name__ == "__main__":
    # read patients list
    # patlist = scipy.io.loadmat('patlist_check.mat')
    # print(patlist['mri_data'])

    """run algorithm for all patients and save the results in patient_res file"""
    # """
    # open patient_res file for writing
    patient_res_wb = load_workbook("patient_res.xlsx")
    patient_res_write = patient_res_wb["Data"]
    row = 2

    # read patient_info file
    patient_info = pd.read_excel("patient_info.xlsx", sheet_name="Data", index_col="patient")
    # print(patient_info)

    # fig = plt.figure()
    # fig.suptitle("inclines inside strip of 9 pixels around the tumor (without registration)")
    # i = 0

    for dir in sorted(os.listdir("patients"), key=lambda x: int(x.split('P')[1])):
        patient_data = [dir, patient_info.at[dir, 'pathological response'],
                        patient_info.at[dir, 'radiological response']]
        # i += 1
        # fig.add_subplot(6, 6, i)


        # incline inside the tumor
        patient_data.append(algorithm(patient=dir, strip_thickness=0, show_images=False))

        # incline inside strips around the tumor
        for strip in range(3, 10, 2):
            patient_data.append(algorithm(patient=dir, strip_thickness=strip, show_images=False))

        # incline in area without tumor
        patient_data.append(algorithm(patient=dir, strip_thickness=0, show_images=False, tumorArea=False,
                                      rows_start=int(patient_info.at[dir, 'rows start black area']),
                                      rows_end=int(patient_info.at[dir, 'rows end black area']),
                                      columns_start=int(patient_info.at[dir, 'columns start black area']),
                                      columns_end=int(patient_info.at[dir, 'columns end black area'])))

        # print(patient_data)
        # for i in range(len(patient_data)):
        #     patient_res_write.cell(row, i + 1).value = patient_data[i]
        # row += 1

    # plt.tight_layout(pad=1, w_pad=0, h_pad=-2)
    # plt.show()

    patient_res_wb.save("patient_res.xlsx")
    # """

    """make analysis on the results in patient_res file"""
    # calculate_average_and_std_of_incline()

    # print_graph_of_average_and_std_of_incline()

    # calculate_difference_in_response_to_treatment()
