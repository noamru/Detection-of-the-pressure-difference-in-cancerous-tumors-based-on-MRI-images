import pandas as pd
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

"""calculate for each kind of strip (3,5,7,9) the average and std of the incline from LinearRegression
           for the groups:
             1. patients that the pathological response was successful.
             2. patients that the pathological response was unsuccessful.
             3. patients that the radiological response was successful.
             4. patients that the radiological response was unsuccessful.
           and save the results in patient_res file"""
def calculate_average_and_std_of_incline():
    # open patient_res file for writing
    patient_res_wb = load_workbook("patient_res.xlsx")
    patient_res_write = patient_res_wb["Data"]

    # read patient_res file for calculations
    patient_res_read = pd.read_excel("patient_res.xlsx", sheet_name="Data")

    write_column = 12

    for column in range(4, 8):  # run on all strips
        strip = patient_res_read.columns[column]

        pathological_successful_treatment_inclines = []
        pathological_unsuccessful_treatment_inclines = []
        radiological_successful_treatment_inclines = []
        radiological_unsuccessful_treatment_inclines = []

        for patient in range(len(patient_res_read)):
            if patient_res_read.at[patient, "pathological response"] == 0 or patient_res_read.at[patient, "pathological response"] == 1:
                pathological_successful_treatment_inclines.append(patient_res_read.at[patient, strip])
            elif patient_res_read.at[patient, "pathological response"] == 2:
                pathological_unsuccessful_treatment_inclines.append(patient_res_read.at[patient, strip])

            if patient_res_read.at[patient, "radiological response"] == 0 or patient_res_read.at[patient, "radiological response"] == 1:
                radiological_successful_treatment_inclines.append(patient_res_read.at[patient, strip])
            elif patient_res_read.at[patient, "radiological response"] == 2:
                radiological_unsuccessful_treatment_inclines.append(patient_res_read.at[patient, strip])

        print(strip, ": ")
        print("pathological: successful-", pathological_successful_treatment_inclines, "unsuccessful-",
              pathological_unsuccessful_treatment_inclines)
        print("radiological: successful-", radiological_successful_treatment_inclines, "unsuccessful-",
              radiological_unsuccessful_treatment_inclines)

        # write in pathological response - successful treatment
        patient_res_write.cell(6, write_column).value = np.mean(pathological_successful_treatment_inclines, axis=0)
        patient_res_write.cell(7, write_column).value = np.std(pathological_successful_treatment_inclines, axis=0)
        write_column += 1

        # write in pathological response - unsuccessful treatment
        patient_res_write.cell(6, write_column).value = np.mean(pathological_unsuccessful_treatment_inclines, axis=0)
        patient_res_write.cell(7, write_column).value = np.std(pathological_unsuccessful_treatment_inclines, axis=0)
        write_column += 1

        # write in radiological response - successful treatment
        patient_res_write.cell(6, write_column).value = np.mean(radiological_successful_treatment_inclines, axis=0)
        patient_res_write.cell(7, write_column).value = np.std(radiological_successful_treatment_inclines, axis=0)
        write_column += 1

        # write in radiological response - unsuccessful treatment
        patient_res_write.cell(6, write_column).value = np.mean(radiological_unsuccessful_treatment_inclines, axis=0)
        patient_res_write.cell(7, write_column).value = np.std(radiological_unsuccessful_treatment_inclines, axis=0)
        write_column += 1

        patient_res_wb.save("patient_res.xlsx")

    print("calculate_average_and_std_of_incline complete\n")


"""show graph of average and std of incline of 3 pixels strip"""
def print_graph_of_average_and_std_of_incline():
    # read patient_res file for calculations
    patient_res_read = pd.read_excel("patient_res.xlsx", sheet_name="Data")

    x = ["pathological response-\nSuccessful treatment", "pathological response-\nUnsuccessful treatment",
         "radiological response-\nSuccessful treatment", "radiological response-\nUnsuccessful treatment"]

    average = [patient_res_read.iat[4, 11], patient_res_read.iat[4, 12], patient_res_read.iat[4, 13], patient_res_read.iat[4, 14]]
    std = [patient_res_read.iat[5, 11], patient_res_read.iat[5, 12], patient_res_read.iat[5, 13], patient_res_read.iat[5, 14]]

    plt.style.use('seaborn-whitegrid')
    plt.errorbar(x, average, std, fmt='o', capsize=5)

    plt.show()



"""Examining whether in cases where there has no leakage from the tumor
    (when the incline in the strip compared to the incline of a non-cancerous area is less than the threshold value)
    if there is a difference in response to treatment between cases that had an increase inside the tumor
    and cases that did not.
    and save the results in patient_res file"""
def calculate_difference_in_response_to_treatment():
    # open patient_res file for writing
    patient_res_wb = load_workbook("patient_res.xlsx")
    patient_res_write = patient_res_wb["Data"]

    # read patient_res file for calculations
    patient_res_read = pd.read_excel("patient_res.xlsx", sheet_name="Data")

    write_column = 30
    for column in range(4, 8):  # run on all strips
        strip = patient_res_read.columns[column]
        write_row = 5
        # strip = patient_res_read.columns[4]  # strip of 3 pixels
        leakage_threshold = 2  # the threshold value for leakage from the tumor
        for patient in range(len(patient_res_read)):
            if patient_res_read.at[patient, strip] - patient_res_read.at[patient, "incline in non-cancerous area"] < leakage_threshold:
                patient_res_write.cell(write_row, write_column).value = patient_res_read.at[patient, "patient"]
                patient_res_write.cell(write_row, write_column + 1).value = str(
                    patient_res_read.at[patient, "incline inside tumor"] > 0)
                patient_res_write.cell(write_row, write_column + 3).value = patient_res_read.at[
                    patient, "pathological response"]
                patient_res_write.cell(write_row, write_column + 5).value = patient_res_read.at[
                    patient, "radiological response"]
                write_row += 1
        write_column += 7

    patient_res_wb.save("patient_res.xlsx")

    print("calculate_difference_in_response_to_treatment complete\n")
